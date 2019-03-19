import openpyxl

#load work book
wk = openpyxl.load_workbook('E:/FreeCRM TestData.xlsx')

print(wk.sheetnames)
print("Active sheet is " + wk.active.title)

#Create object of any sheet on which you want to work
#print sheet name
sh = wk['Contacts']
print(sh.title)

#reading the sheet values
sh = wk["Contacts"]
print(sh['A3'].value)
print(sh['B4'].value)

#printing column names
c1 = sh.cell(4,2)
print(c1.value)

print(c1.row)
print(c1.column)

#Find Rows having data

rows = sh.max_row
columns = sh.max_column

print("Total Rows are-" + str(rows))
print("Total Column are-" + str(columns))

#for reading all data from excel
for i in range(1,rows+1):
    for j in range(1, columns+1):
        c = sh.cell(i,j)
        print(c.value)

#Another method
#for r in sh['A1':'D4']:
    #for c in r:
        #print(c.value)

